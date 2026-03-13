from __future__ import annotations

import csv
import hashlib
import io
import os
import re
import sqlite3
import uuid
from dataclasses import dataclass
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
import smtplib
from typing import Annotated, Literal

from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field
from openpyxl import Workbook, load_workbook

try:
    import psycopg
    from psycopg.rows import dict_row
except Exception:  # pragma: no cover
    psycopg = None
    dict_row = None

BASE_DIR = Path(__file__).parent
STORAGE_DIR = BASE_DIR / "storage"
REPORTS_DIR = STORAGE_DIR / "reports"
DB_PATH = STORAGE_DIR / "fitness.db"
DATABASE_URL = os.getenv("DATABASE_URL")
USE_POSTGRES = bool(DATABASE_URL and DATABASE_URL.startswith("postgres"))

REPORTS_DIR.mkdir(parents=True, exist_ok=True)

app = FastAPI(title="Fitness Test Submission App", version="1.0.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")


class SoldierInput(BaseModel):
    name: str = Field(min_length=1)
    detail_number: str = Field(min_length=1)
    wbt: float
    rir: float
    mcs_level: int
    mcs_shuttle: int


class SubmissionInput(BaseModel):
    unit: str = Field(min_length=1)
    coy: str = Field(min_length=1)
    platoon: str = Field(min_length=1)
    session_code: str = Field(min_length=1)
    test_date: str = Field(min_length=1)
    submitted_by: str = Field(min_length=1)
    notification_email: str | None = None
    soldiers: Annotated[list[SoldierInput], Field(min_length=1)]


class ConductingSessionInput(BaseModel):
    unit: str = Field(min_length=1)
    coy: str = Field(min_length=1)
    test_date: str = Field(min_length=1)
    session_code: str = Field(min_length=1)
    password: str = Field(min_length=1)


class SessionCodeInput(BaseModel):
    session_code: str = Field(min_length=1)


class SoldierProfileInput(BaseModel):
    session_code: str = Field(min_length=1)
    full_nric: str = Field(min_length=1)
    full_name: str = Field(min_length=1)
    rank: str = Field(min_length=1)
    unit: str = Field(min_length=1)
    coy: str = Field(min_length=1)
    platoon: str = Field(min_length=1)
    detail_level: int


ALLOWED_RANKS = {
    "REC",
    "PTE",
    "LCP",
    "CPL",
    "3SG",
    "2SG",
    "1SG",
    "SSG",
    "MSG",
    "3WO",
    "2WO",
    "1WO",
    "OCT",
    "SCT",
    "2LT",
    "LTA",
    "CPT",
    "MAJ",
    "LTC",
    "SLTC",
    "COL",
}


class CommanderLoginInput(BaseModel):
    session_code: str = Field(min_length=1)
    password: str = Field(min_length=1)
    test_date: str = Field(min_length=1)


class CommanderScoreRow(BaseModel):
    soldier_id: str = Field(min_length=1)
    wbt: float
    rir: float
    mcs_stage: int
    mcs_level: int


class CommanderSaveScoresInput(CommanderLoginInput):
    scores: Annotated[list[CommanderScoreRow], Field(min_length=1)]


class CommanderStationScoreRow(BaseModel):
    soldier_id: str = Field(min_length=1)
    wbt: str | None = None
    rir: str | None = None
    mcs_stage: str | None = None
    mcs_level: str | None = None


class CommanderStationSaveInput(CommanderLoginInput):
    station: Literal["WBT", "RIR", "MCS"]
    scores: Annotated[list[CommanderStationScoreRow], Field(min_length=1)]


class OfficerSessionAuthInput(BaseModel):
    session_code: str = Field(min_length=1)
    password: str = Field(min_length=1)
    test_date: str = Field(min_length=1)


@dataclass
class EmailResult:
    sent: bool
    message: str


class DBConnection:
    def __init__(self) -> None:
        if USE_POSTGRES:
            if psycopg is None:
                raise RuntimeError("PostgreSQL backend requested but psycopg is not installed")
            self._conn = psycopg.connect(DATABASE_URL, row_factory=dict_row)
            self._is_postgres = True
        else:
            self._conn = sqlite3.connect(DB_PATH)
            self._conn.row_factory = sqlite3.Row
            self._is_postgres = False

    def _adapt_query(self, query: str) -> str:
        if not self._is_postgres:
            return query
        return query.replace("?", "%s")

    def execute(self, query: str, params: tuple | list = ()):
        cursor = self._conn.cursor()
        cursor.execute(self._adapt_query(query), params)
        return cursor

    def executemany(self, query: str, params_seq: list[tuple]):
        cursor = self._conn.cursor()
        cursor.executemany(self._adapt_query(query), params_seq)
        return cursor

    def commit(self) -> None:
        self._conn.commit()

    def rollback(self) -> None:
        self._conn.rollback()

    def close(self) -> None:
        self._conn.close()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc, tb):
        if exc_type is None:
            self.commit()
        else:
            self.rollback()
        self.close()


def get_conn() -> DBConnection:
    return DBConnection()


def init_db() -> None:
    with get_conn() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS conducting_sessions (
                session_id TEXT PRIMARY KEY,
                unit TEXT NOT NULL,
                coy TEXT NOT NULL,
                test_date TEXT NOT NULL,
                session_code TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(session_code, password_hash, test_date)
            )
            """
        )
        if not USE_POSTGRES:
            create_sql_row = conn.execute(
                "SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'conducting_sessions'",
            ).fetchone()
            create_sql = (create_sql_row["sql"] if create_sql_row else "") or ""
            if "session_code TEXT NOT NULL UNIQUE" in create_sql:
                conn.execute("PRAGMA foreign_keys=OFF")
                conn.execute(
                    """
                    CREATE TABLE IF NOT EXISTS conducting_sessions_new (
                        session_id TEXT PRIMARY KEY,
                        unit TEXT NOT NULL,
                        coy TEXT NOT NULL,
                        test_date TEXT NOT NULL,
                        session_code TEXT NOT NULL,
                        password_hash TEXT NOT NULL,
                        created_at TEXT NOT NULL,
                        UNIQUE(session_code, password_hash, test_date)
                    )
                    """
                )
                conn.execute(
                    """
                    INSERT INTO conducting_sessions_new (
                        session_id, unit, coy, test_date, session_code, password_hash, created_at
                    )
                    SELECT session_id, unit, coy, test_date, session_code, password_hash, created_at
                    FROM conducting_sessions
                    """
                )
                conn.execute("DROP TABLE conducting_sessions")
                conn.execute("ALTER TABLE conducting_sessions_new RENAME TO conducting_sessions")
                conn.execute("PRAGMA foreign_keys=ON")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS soldier_profiles (
                soldier_id TEXT PRIMARY KEY,
                session_id TEXT NOT NULL,
                full_nric TEXT NOT NULL,
                full_name TEXT NOT NULL,
                rank TEXT NOT NULL DEFAULT '',
                unit TEXT NOT NULL,
                coy TEXT NOT NULL,
                platoon TEXT NOT NULL,
                detail_level TEXT NOT NULL,
                created_at TEXT NOT NULL,
                UNIQUE(session_id, full_nric),
                FOREIGN KEY (session_id) REFERENCES conducting_sessions(session_id)
            )
            """
        )
        if not USE_POSTGRES:
            existing_columns = {
                row["name"]
                for row in conn.execute("PRAGMA table_info(soldier_profiles)").fetchall()
            }
            if "rank" not in existing_columns:
                conn.execute("ALTER TABLE soldier_profiles ADD COLUMN rank TEXT NOT NULL DEFAULT ''")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS soldier_test_scores (
                score_id TEXT PRIMARY KEY,
                soldier_id TEXT NOT NULL UNIQUE,
                wbt TEXT,
                rir TEXT,
                mcs_stage TEXT,
                mcs_level TEXT,
                updated_at TEXT NOT NULL,
                FOREIGN KEY (soldier_id) REFERENCES soldier_profiles(soldier_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS commander_exports (
                export_id TEXT PRIMARY KEY,
                session_id TEXT NOT NULL,
                report_path TEXT NOT NULL,
                share_link TEXT NOT NULL,
                created_at TEXT NOT NULL,
                FOREIGN KEY (session_id) REFERENCES conducting_sessions(session_id)
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS submission_sessions (
                submission_id TEXT PRIMARY KEY,
                unit TEXT NOT NULL,
                coy TEXT NOT NULL,
                platoon TEXT NOT NULL,
                session_code TEXT NOT NULL,
                test_date TEXT NOT NULL,
                submitted_by TEXT NOT NULL,
                submitted_at TEXT NOT NULL,
                report_path TEXT NOT NULL,
                share_link TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS soldier_scores (
                score_id TEXT PRIMARY KEY,
                submission_id TEXT NOT NULL,
                name TEXT NOT NULL,
                detail_number TEXT NOT NULL,
                wbt REAL NOT NULL,
                rir REAL NOT NULL,
                mcs_level INTEGER NOT NULL,
                mcs_shuttle INTEGER NOT NULL,
                FOREIGN KEY (submission_id) REFERENCES submission_sessions(submission_id)
            )
            """
        )


@app.on_event("startup")
def startup() -> None:
    init_db()


@app.get("/")
def root() -> FileResponse:
    return FileResponse(BASE_DIR / "static" / "index.html")


@app.get("/conducting-officer")
def conducting_officer_page() -> FileResponse:
    return FileResponse(BASE_DIR / "static" / "conducting.html")


@app.get("/conducting-officer/dashboard")
def conducting_officer_dashboard_page() -> FileResponse:
    return FileResponse(BASE_DIR / "static" / "conducting_dashboard.html")


@app.get("/soldier/login")
def soldier_login_page() -> FileResponse:
    return FileResponse(BASE_DIR / "static" / "soldier_login.html")


@app.get("/soldier/details")
def soldier_details_page() -> FileResponse:
    return FileResponse(BASE_DIR / "static" / "soldier_details.html")


@app.get("/commander/login")
def commander_login_page() -> FileResponse:
    return FileResponse(BASE_DIR / "static" / "commander_login.html")


@app.get("/commander/dashboard")
def commander_dashboard_page() -> FileResponse:
    return FileResponse(BASE_DIR / "static" / "commander_dashboard.html")


@app.get("/api/meta/options")
def get_options() -> JSONResponse:
    data = {
        "units": ["1 SIR", "2 SIR", "HQ"],
        "coys": ["Alpha", "Bravo", "Charlie", "Support"],
        "platoons": ["Platoon 1", "Platoon 2", "Platoon 3", "Platoon 4"],
    }
    return JSONResponse(data)


def hash_password(raw_password: str) -> str:
    return hashlib.sha256(raw_password.encode("utf-8")).hexdigest()


MCS_STAGE_RANGE_BY_LEVEL: dict[int, tuple[int, int]] = {
    1: (0, 0),
    2: (0, 0),
    3: (0, 0),
    4: (1, 1),
    5: (1, 2),
    6: (1, 2),
    7: (1, 4),
    8: (1, 6),
    9: (1, 8),
    10: (1, 8),
    11: (1, 10),
    12: (1, 12),
    13: (1, 14),
    14: (1, 12),
    15: (1, 9),
    16: (1, 2),
}


def normalize_station_score(value: str | None, min_value: int, max_value: int, label: str) -> str:
    if value is None:
        raise HTTPException(status_code=400, detail=f"{label} value is required")

    cleaned = value.strip().upper()
    if cleaned == "DNF":
        return "DNF"

    if not cleaned.isdigit():
        raise HTTPException(status_code=400, detail=f"{label} must be an integer between {min_value} and {max_value} or DNF")

    parsed = int(cleaned)
    if parsed < min_value or parsed > max_value:
        raise HTTPException(status_code=400, detail=f"{label} must be between {min_value} and {max_value} or DNF")

    return str(parsed)


def normalize_mcs_stage_by_level(level_value: str, stage_value: str | None) -> str:
    if stage_value is None:
        raise HTTPException(status_code=400, detail="MCS Stage value is required")

    cleaned_stage = stage_value.strip().upper()
    if cleaned_stage == "DNF":
        return "DNF"

    if level_value == "DNF":
        raise HTTPException(status_code=400, detail="MCS Stage must be DNF when MCS Level is DNF")

    if not cleaned_stage.isdigit():
        raise HTTPException(status_code=400, detail="MCS Stage must be an integer in the valid range or DNF")

    level_int = int(level_value)
    stage_int = int(cleaned_stage)
    stage_range = MCS_STAGE_RANGE_BY_LEVEL.get(level_int)

    if stage_range is None:
        raise HTTPException(status_code=400, detail="MCS Level is invalid")

    min_stage, max_stage = stage_range
    if stage_int < min_stage or stage_int > max_stage:
        raise HTTPException(
            status_code=400,
            detail=f"MCS Stage must be between {min_stage} and {max_stage} for MCS Level {level_int}, or DNF",
        )

    return str(stage_int)


def get_session_by_code(session_code: str) -> sqlite3.Row | None:
    with get_conn() as conn:
        return conn.execute(
            """
            SELECT * FROM conducting_sessions
            WHERE session_code = ?
            ORDER BY test_date DESC, created_at DESC
            LIMIT 1
            """,
            (session_code.strip().upper(),),
        ).fetchone()


def get_session_by_credentials(session_code: str, password: str, test_date: str) -> sqlite3.Row | None:
    password_hash = hash_password(password)
    with get_conn() as conn:
        return conn.execute(
            """
            SELECT * FROM conducting_sessions
            WHERE session_code = ? AND password_hash = ? AND test_date = ?
            LIMIT 1
            """,
            (session_code.strip().upper(), password_hash, test_date.strip()),
        ).fetchone()


def get_commander_session_payload(session_row: sqlite3.Row) -> dict:
    with get_conn() as conn:
        soldiers = conn.execute(
            """
            SELECT
                sp.soldier_id,
                sp.full_nric,
                sp.full_name,
                sp.rank,
                sp.unit,
                sp.coy,
                sp.platoon,
                sp.detail_level,
                sts.wbt,
                sts.rir,
                sts.mcs_stage,
                sts.mcs_level
            FROM soldier_profiles sp
            LEFT JOIN soldier_test_scores sts ON sp.soldier_id = sts.soldier_id
            WHERE sp.session_id = ?
            ORDER BY sp.detail_level, sp.full_name
            """,
            (session_row["session_id"],),
        ).fetchall()

    grouped: dict[str, list[dict]] = {}
    for soldier in soldiers:
        detail_level = soldier["detail_level"]
        grouped.setdefault(detail_level, []).append(dict(soldier))

    return {
        "session": {
            "session_code": session_row["session_code"],
            "unit": session_row["unit"],
            "coy": session_row["coy"],
            "test_date": session_row["test_date"],
        },
        "by_detail_level": grouped,
    }


def build_commander_export_csv(export_id: str, session_row: sqlite3.Row) -> Path:
    report_file = REPORTS_DIR / f"commander_{export_id}.csv"
    fieldnames = [
        "TEST_DATE",
        "NRIC",
        "PLATOON",
        "UNIT",
        "COY",
        "RANK",
        "NAME",
        "WBT",
        "RIR",
        "MCS_LEVEL",
        "MCS_STAGE",
    ]

    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT
                cs.session_code,
                cs.test_date,
                sp.unit,
                sp.coy,
                sp.platoon,
                sp.detail_level,
                sp.rank,
                sp.full_nric,
                sp.full_name,
                sts.wbt,
                sts.rir,
                sts.mcs_level,
                sts.mcs_stage
            FROM conducting_sessions cs
            JOIN soldier_profiles sp ON cs.session_id = sp.session_id
            LEFT JOIN soldier_test_scores sts ON sp.soldier_id = sts.soldier_id
            WHERE cs.session_id = ?
            ORDER BY sp.detail_level, sp.full_name
            """,
            (session_row["session_id"],),
        ).fetchall()

    with report_file.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "TEST_DATE": row["test_date"],
                    "NRIC": row["full_nric"],
                    "PLATOON": row["platoon"],
                    "UNIT": row["unit"],
                    "COY": row["coy"],
                    "RANK": row["rank"],
                    "NAME": row["full_name"],
                    "WBT": row["wbt"],
                    "RIR": row["rir"],
                    "MCS_LEVEL": row["mcs_level"],
                    "MCS_STAGE": row["mcs_stage"],
                }
            )

    return report_file


def normalize_filename_part(value: str | None) -> str:
    normalized = re.sub(r"[^A-Z0-9-]+", "_", str(value or "").strip().upper())
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized or "NA"


def build_officer_export_download_name(unit: str | None, coy: str | None, test_date: str | None) -> str:
    safe_unit = normalize_filename_part(unit)
    safe_coy = normalize_filename_part(coy)
    safe_date = normalize_filename_part(test_date)
    return f"{safe_unit}_{safe_coy}_{safe_date}_ACFC.csv"


DETAIL_IMPORT_REQUIRED_HEADERS = {
    "NRIC",
    "NAME",
    "RANK",
    "UNIT",
    "COY",
    "PLATOON",
    "DETAIL_LEVEL",
}


def normalize_import_header(value: str | None) -> str:
    cleaned = str(value or "").strip().upper()
    cleaned = re.sub(r"[^A-Z0-9]+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    if cleaned == "FULL_NRIC":
        return "NRIC"
    if cleaned == "FULL_NAME":
        return "NAME"
    return cleaned


def parse_detail_level(value: str, row_number: int) -> str:
    cleaned = value.strip()
    if not cleaned.isdigit():
        raise HTTPException(status_code=400, detail=f"Row {row_number}: DETAIL_LEVEL must be an integer between 1 and 20")
    parsed = int(cleaned)
    if parsed < 1 or parsed > 20:
        raise HTTPException(status_code=400, detail=f"Row {row_number}: DETAIL_LEVEL must be between 1 and 20")
    return str(parsed)


def clear_session_profiles(conn: DBConnection, session_id: str) -> tuple[int, int]:
    soldier_rows = conn.execute(
        "SELECT soldier_id FROM soldier_profiles WHERE session_id = ?",
        (session_id,),
    ).fetchall()
    soldier_ids = [row["soldier_id"] for row in soldier_rows]

    deleted_scores = 0
    if soldier_ids:
        placeholders = ",".join("?" for _ in soldier_ids)
        deleted_scores = conn.execute(
            f"DELETE FROM soldier_test_scores WHERE soldier_id IN ({placeholders})",
            tuple(soldier_ids),
        ).rowcount

    deleted_profiles = conn.execute(
        "DELETE FROM soldier_profiles WHERE session_id = ?",
        (session_id,),
    ).rowcount

    return deleted_profiles or 0, deleted_scores or 0


@app.get("/api/officer/import-template")
def download_detail_import_template() -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "DETAIL_SHEET"

    headers = ["NRIC", "NAME", "RANK", "UNIT", "COY", "PLATOON", "DETAIL_LEVEL"]
    sheet.append(headers)
    sheet.append(["S1234567A", "TAN AH KOW", "CPL", "1 SIR", "ALPHA", "1", "4"])

    content = io.BytesIO()
    workbook.save(content)
    content.seek(0)

    return StreamingResponse(
        content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=ACFC_DETAIL_IMPORT_TEMPLATE.xlsx"},
    )


@app.post("/api/officer/import-details")
async def import_officer_details(
    session_code: str = Form(...),
    password: str = Form(...),
    test_date: str = Form(...),
    file: UploadFile = File(...),
) -> JSONResponse:
    session_code_value = session_code.strip().upper()
    password_value = password
    test_date_value = test_date.strip()

    session_row = get_session_by_credentials(session_code_value, password_value, test_date_value)
    if not session_row:
        raise HTTPException(status_code=401, detail="Invalid Session Code, Password, or Test Date")

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Upload must be an .xlsx file")

    with get_conn() as conn:
        existing_count = conn.execute(
            "SELECT COUNT(1) AS total FROM soldier_profiles WHERE session_id = ?",
            (session_row["session_id"],),
        ).fetchone()["total"]
        if int(existing_count or 0) > 0:
            raise HTTPException(
                status_code=409,
                detail="This session already has imported details. Clear imported details before uploading a new file.",
            )

    file_bytes = await file.read()
    try:
        workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Unable to read Excel file. Please use the provided template.")

    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    raw_headers = rows[0]
    normalized_headers = [normalize_import_header(h) for h in raw_headers]
    missing_headers = [header for header in DETAIL_IMPORT_REQUIRED_HEADERS if header not in normalized_headers]
    if missing_headers:
        raise HTTPException(status_code=400, detail=f"Missing required columns: {', '.join(sorted(missing_headers))}")

    index_by_header = {header: idx for idx, header in enumerate(normalized_headers)}

    prepared_rows: list[tuple] = []
    seen_nric: set[str] = set()
    created_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"

    for index, excel_row in enumerate(rows[1:], start=2):
        if excel_row is None:
            continue

        def value_for(header: str) -> str:
            position = index_by_header[header]
            if position >= len(excel_row):
                return ""
            return str(excel_row[position] or "").strip().upper()

        nric = value_for("NRIC")
        name = value_for("NAME")
        rank = value_for("RANK")
        unit = value_for("UNIT")
        coy = value_for("COY")
        platoon = value_for("PLATOON")
        detail_level_raw = value_for("DETAIL_LEVEL")

        if not any([nric, name, rank, unit, coy, platoon, detail_level_raw]):
            continue

        if not nric or not name or not rank or not unit or not coy or not platoon or not detail_level_raw:
            raise HTTPException(status_code=400, detail=f"Row {index}: all required columns must be filled")

        if rank not in ALLOWED_RANKS:
            raise HTTPException(status_code=400, detail=f"Row {index}: RANK '{rank}' is invalid")

        detail_level = parse_detail_level(detail_level_raw, index)

        if nric in seen_nric:
            raise HTTPException(status_code=400, detail=f"Row {index}: duplicate NRIC '{nric}' in file")
        seen_nric.add(nric)

        prepared_rows.append(
            (
                str(uuid.uuid4()),
                session_row["session_id"],
                nric,
                name,
                rank,
                unit,
                coy,
                platoon,
                detail_level,
                created_at,
            )
        )

    if not prepared_rows:
        raise HTTPException(status_code=400, detail="No valid data rows found in uploaded file")

    try:
        with get_conn() as conn:
            conn.executemany(
                """
                INSERT INTO soldier_profiles (
                    soldier_id, session_id, full_nric, full_name, rank, unit,
                    coy, platoon, detail_level, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                prepared_rows,
            )
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Import failed due to duplicate NRIC in this session")

    return JSONResponse(
        {
            "session_code": session_code_value,
            "imported_count": len(prepared_rows),
            "message": "Detail sheet imported successfully.",
        }
    )


@app.post("/api/officer/import-details/clear")
def clear_officer_imported_details(payload: OfficerSessionAuthInput) -> JSONResponse:
    session_row = get_session_by_credentials(payload.session_code, payload.password, payload.test_date)
    if not session_row:
        raise HTTPException(status_code=401, detail="Invalid Session Code, Password, or Test Date")

    with get_conn() as conn:
        deleted_profiles, deleted_scores = clear_session_profiles(conn, session_row["session_id"])

    return JSONResponse(
        {
            "session_code": session_row["session_code"],
            "deleted_profiles": deleted_profiles,
            "deleted_scores": deleted_scores,
            "message": "Imported detail data cleared. You can upload a new file now.",
        }
    )


@app.post("/api/conducting/sessions")
def create_conducting_session(payload: ConductingSessionInput) -> JSONResponse:
    session_id = str(uuid.uuid4())
    created_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    password_hash = hash_password(payload.password)
    session_code_value = payload.session_code.strip().upper()
    unit_value = payload.unit.strip().upper()
    coy_value = payload.coy.strip().upper()

    with get_conn() as conn:
        conduct_exists = conn.execute(
            "SELECT 1 FROM conducting_sessions WHERE session_code = ? AND password_hash = ? AND test_date = ?",
            (session_code_value, password_hash, payload.test_date),
        ).fetchone()

        if conduct_exists:
            raise HTTPException(
                status_code=409,
                detail="A conduct with the same Session Code, Password, and Test Date already exists.",
            )

        try:
            conn.execute(
                """
                INSERT INTO conducting_sessions (
                    session_id, unit, coy, test_date, session_code, password_hash, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    session_id,
                    unit_value,
                    coy_value,
                    payload.test_date,
                    session_code_value,
                    password_hash,
                    created_at,
                ),
            )
        except sqlite3.IntegrityError:
            raise HTTPException(
                status_code=409,
                detail="A conduct with the same Session Code, Password, and Test Date already exists.",
            )

    return JSONResponse(
        {
            "session_id": session_id,
            "session_code": session_code_value,
            "created_at": created_at,
            "message": "Session created successfully.",
        }
    )


@app.post("/api/soldier/session/validate")
def validate_soldier_session(payload: SessionCodeInput) -> JSONResponse:
    row = get_session_by_code(payload.session_code)
    if not row:
        raise HTTPException(status_code=404, detail="Session code not found")

    return JSONResponse(
        {
            "session_code": row["session_code"],
            "unit": row["unit"],
            "coy": row["coy"],
            "test_date": row["test_date"],
        }
    )


@app.post("/api/soldier/profiles")
def create_soldier_profile(payload: SoldierProfileInput) -> JSONResponse:
    session_row = get_session_by_code(payload.session_code)
    if not session_row:
        raise HTTPException(status_code=404, detail="Session code not found")

    if payload.detail_level < 1 or payload.detail_level > 20:
        raise HTTPException(status_code=400, detail="Detail level must be an integer between 1 and 20")

    rank_value = payload.rank.strip().upper()
    if rank_value not in ALLOWED_RANKS:
        raise HTTPException(status_code=400, detail="Rank is invalid")

    soldier_id = str(uuid.uuid4())
    created_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"

    try:
        with get_conn() as conn:
            conn.execute(
                """
                INSERT INTO soldier_profiles (
                    soldier_id, session_id, full_nric, full_name, rank, unit,
                    coy, platoon, detail_level, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    soldier_id,
                    session_row["session_id"],
                    payload.full_nric,
                    payload.full_name,
                    rank_value,
                    payload.unit,
                    payload.coy,
                    payload.platoon,
                    str(payload.detail_level),
                    created_at,
                ),
            )
    except sqlite3.IntegrityError:
        raise HTTPException(status_code=409, detail="Soldier already registered for this session")

    return JSONResponse(
        {
            "soldier_id": soldier_id,
            "session_code": payload.session_code,
            "message": "Soldier details submitted successfully.",
        }
    )


@app.post("/api/commander/login")
def commander_login(payload: CommanderLoginInput) -> JSONResponse:
    row = get_session_by_credentials(payload.session_code, payload.password, payload.test_date)
    if not row:
        raise HTTPException(status_code=401, detail="Invalid Session Code, Password, or Test Date")

    return JSONResponse(
        {
            "session_code": row["session_code"],
            "unit": row["unit"],
            "coy": row["coy"],
            "test_date": row["test_date"],
            "message": "Login successful.",
        }
    )


@app.post("/api/commander/session-data")
def commander_session_data(payload: CommanderLoginInput) -> JSONResponse:
    row = get_session_by_credentials(payload.session_code, payload.password, payload.test_date)
    if not row:
        raise HTTPException(status_code=401, detail="Invalid Session Code, Password, or Test Date")
    return JSONResponse(get_commander_session_payload(row))


@app.post("/api/officer/session-data")
def officer_session_data(payload: CommanderLoginInput) -> JSONResponse:
    row = get_session_by_credentials(payload.session_code, payload.password, payload.test_date)
    if not row:
        raise HTTPException(status_code=401, detail="Invalid Session Code, Password, or Test Date")
    return JSONResponse(get_commander_session_payload(row))


@app.post("/api/commander/scores")
def save_commander_scores(payload: CommanderSaveScoresInput) -> JSONResponse:
    row = get_session_by_credentials(payload.session_code, payload.password, payload.test_date)
    if not row:
        raise HTTPException(status_code=401, detail="Invalid Session Code, Password, or Test Date")

    updated_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"

    with get_conn() as conn:
        owned_soldiers = conn.execute(
            "SELECT soldier_id FROM soldier_profiles WHERE session_id = ?",
            (row["session_id"],),
        ).fetchall()
        owned_ids = {r["soldier_id"] for r in owned_soldiers}

        for score in payload.scores:
            if score.soldier_id not in owned_ids:
                raise HTTPException(status_code=400, detail="Score row includes soldier outside this session")

            conn.execute(
                """
                INSERT INTO soldier_test_scores (
                    score_id, soldier_id, wbt, rir, mcs_stage, mcs_level, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(soldier_id) DO UPDATE SET
                    wbt = excluded.wbt,
                    rir = excluded.rir,
                    mcs_stage = excluded.mcs_stage,
                    mcs_level = excluded.mcs_level,
                    updated_at = excluded.updated_at
                """,
                (
                    str(uuid.uuid4()),
                    score.soldier_id,
                    score.wbt,
                    score.rir,
                    score.mcs_stage,
                    score.mcs_level,
                    updated_at,
                ),
            )

    return JSONResponse(
        {
            "updated_count": len(payload.scores),
            "updated_at": updated_at,
            "message": "Scores saved successfully.",
        }
    )


@app.post("/api/commander/scores/station")
def save_commander_station_scores(payload: CommanderStationSaveInput) -> JSONResponse:
    row = get_session_by_credentials(payload.session_code, payload.password, payload.test_date)
    if not row:
        raise HTTPException(status_code=401, detail="Invalid Session Code, Password, or Test Date")

    updated_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"

    with get_conn() as conn:
        owned_soldiers = conn.execute(
            "SELECT soldier_id FROM soldier_profiles WHERE session_id = ?",
            (row["session_id"],),
        ).fetchall()
        owned_ids = {r["soldier_id"] for r in owned_soldiers}

        for score in payload.scores:
            if score.soldier_id not in owned_ids:
                raise HTTPException(status_code=400, detail="Score row includes soldier outside this session")

            wbt_value: str | None = None
            rir_value: str | None = None
            mcs_stage_value: str | None = None
            mcs_level_value: str | None = None

            if payload.station == "WBT":
                wbt_value = normalize_station_score(score.wbt, 0, 100, "WBT")
            if payload.station == "RIR":
                rir_value = normalize_station_score(score.rir, 0, 100, "RIR")
            if payload.station == "MCS":
                mcs_level_value = normalize_station_score(score.mcs_level, 1, 16, "MCS Level")
                mcs_stage_value = normalize_mcs_stage_by_level(mcs_level_value, score.mcs_stage)

            conn.execute(
                """
                INSERT INTO soldier_test_scores (
                    score_id, soldier_id, wbt, rir, mcs_stage, mcs_level, updated_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(soldier_id) DO UPDATE SET
                    wbt = COALESCE(excluded.wbt, soldier_test_scores.wbt),
                    rir = COALESCE(excluded.rir, soldier_test_scores.rir),
                    mcs_stage = COALESCE(excluded.mcs_stage, soldier_test_scores.mcs_stage),
                    mcs_level = COALESCE(excluded.mcs_level, soldier_test_scores.mcs_level),
                    updated_at = excluded.updated_at
                """,
                (
                    str(uuid.uuid4()),
                    score.soldier_id,
                    wbt_value,
                    rir_value,
                    mcs_stage_value,
                    mcs_level_value,
                    updated_at,
                ),
            )

    return JSONResponse(
        {
            "station": payload.station,
            "updated_count": len(payload.scores),
            "updated_at": updated_at,
            "message": f"{payload.station} scores saved successfully.",
        }
    )


@app.post("/api/officer/export")
def export_officer_session(payload: CommanderLoginInput, request: Request) -> JSONResponse:
    session_row = get_session_by_credentials(payload.session_code, payload.password, payload.test_date)
    if not session_row:
        raise HTTPException(status_code=401, detail="Invalid Session Code, Password, or Test Date")

    export_id = str(uuid.uuid4())
    created_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    report_file = build_commander_export_csv(export_id, session_row)
    base_url = os.getenv("APP_BASE_URL") or str(request.base_url).rstrip("/")
    share_link = f"{base_url}/api/officer/reports/{export_id}/download"

    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO commander_exports (export_id, session_id, report_path, share_link, created_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (
                export_id,
                session_row["session_id"],
                str(report_file),
                share_link,
                created_at,
            ),
        )

    return JSONResponse(
        {
            "export_id": export_id,
            "created_at": created_at,
            "report_file": str(report_file.relative_to(BASE_DIR)),
            "share_link": share_link,
            "message": "Officer session CSV generated.",
        }
    )


@app.get("/api/officer/reports/{export_id}/download")
def download_officer_export(export_id: str) -> FileResponse:
    with get_conn() as conn:
        row = conn.execute(
            """
            SELECT ce.report_path, cs.unit, cs.coy, cs.test_date
            FROM commander_exports ce
            JOIN conducting_sessions cs ON cs.session_id = ce.session_id
            WHERE ce.export_id = ?
            """,
            (export_id,),
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Commander export not found")

    report_path = Path(row["report_path"])
    if not report_path.exists():
        raise HTTPException(status_code=404, detail="Stored commander report file is missing")

    download_name = build_officer_export_download_name(
        row["unit"],
        row["coy"],
        row["test_date"],
    )

    return FileResponse(
        report_path,
        media_type="text/csv",
        filename=download_name,
    )


def build_csv_file(submission_id: str, payload: SubmissionInput, submitted_at: str) -> Path:
    report_file = REPORTS_DIR / f"{submission_id}.csv"
    fieldnames = [
        "UNIT",
        "COY",
        "PLATOON",
        "NAME",
        "DETAIL_NUMBER",
        "SESSION_CODE",
        "TEST_DATE",
        "WBT",
        "RIR",
        "MCS_LEVEL",
        "MCS_SHUTTLE",
        "SUBMISSION_ID",
        "SUBMITTED_AT",
    ]

    with report_file.open("w", newline="", encoding="utf-8") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
        writer.writeheader()
        for soldier in payload.soldiers:
            writer.writerow(
                {
                    "UNIT": payload.unit,
                    "COY": payload.coy,
                    "PLATOON": payload.platoon,
                    "NAME": soldier.name,
                    "DETAIL_NUMBER": soldier.detail_number,
                    "SESSION_CODE": payload.session_code,
                    "TEST_DATE": payload.test_date,
                    "WBT": soldier.wbt,
                    "RIR": soldier.rir,
                    "MCS_LEVEL": soldier.mcs_level,
                    "MCS_SHUTTLE": soldier.mcs_shuttle,
                    "SUBMISSION_ID": submission_id,
                    "SUBMITTED_AT": submitted_at,
                }
            )

    return report_file


def try_send_email(recipient: str, file_path: Path, submission_id: str) -> EmailResult:
    smtp_host = os.getenv("SMTP_HOST")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    smtp_user = os.getenv("SMTP_USER")
    smtp_pass = os.getenv("SMTP_PASS")
    smtp_from = os.getenv("SMTP_FROM", smtp_user or "")

    if not smtp_host or not smtp_user or not smtp_pass or not smtp_from:
        return EmailResult(False, "SMTP settings not configured. Skipped email send.")

    message = EmailMessage()
    message["Subject"] = f"Fitness Test Submission {submission_id}"
    message["From"] = smtp_from
    message["To"] = recipient
    message.set_content(f"Attached is the report for submission {submission_id}.")

    with file_path.open("rb") as f:
        file_bytes = f.read()

    message.add_attachment(file_bytes, maintype="text", subtype="csv", filename=file_path.name)

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(message)
        return EmailResult(True, "Email sent successfully.")
    except Exception as exc:
        return EmailResult(False, f"Email send failed: {exc}")


@app.post("/api/submissions")
def submit_scores(payload: SubmissionInput, request: Request) -> JSONResponse:
    submission_id = str(uuid.uuid4())
    submitted_at = datetime.utcnow().isoformat(timespec="seconds") + "Z"

    report_file = build_csv_file(submission_id, payload, submitted_at)
    base_url = os.getenv("APP_BASE_URL") or str(request.base_url).rstrip("/")
    share_link = f"{base_url}/api/reports/{submission_id}/download"

    with get_conn() as conn:
        conn.execute(
            """
            INSERT INTO submission_sessions (
                submission_id, unit, coy, platoon, session_code, test_date,
                submitted_by, submitted_at, report_path, share_link
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                submission_id,
                payload.unit,
                payload.coy,
                payload.platoon,
                payload.session_code,
                payload.test_date,
                payload.submitted_by,
                submitted_at,
                str(report_file),
                share_link,
            ),
        )

        conn.executemany(
            """
            INSERT INTO soldier_scores (
                score_id, submission_id, name, detail_number,
                wbt, rir, mcs_level, mcs_shuttle
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                (
                    str(uuid.uuid4()),
                    submission_id,
                    soldier.name,
                    soldier.detail_number,
                    soldier.wbt,
                    soldier.rir,
                    soldier.mcs_level,
                    soldier.mcs_shuttle,
                )
                for soldier in payload.soldiers
            ],
        )

    email_result = EmailResult(False, "No recipient provided.")
    if payload.notification_email:
        email_result = try_send_email(payload.notification_email, report_file, submission_id)

    return JSONResponse(
        {
            "submission_id": submission_id,
            "submitted_at": submitted_at,
            "soldier_count": len(payload.soldiers),
            "report_file": str(report_file.relative_to(BASE_DIR)),
            "share_link": share_link,
            "email_sent": email_result.sent,
            "email_message": email_result.message,
            "message": "Submission received. CSV generated and stored.",
        }
    )


@app.get("/api/reports/{submission_id}/download")
def download_report(submission_id: str) -> FileResponse:
    with get_conn() as conn:
        row = conn.execute(
            "SELECT report_path FROM submission_sessions WHERE submission_id = ?",
            (submission_id,),
        ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Report not found")

    report_path = Path(row["report_path"])
    if not report_path.exists():
        raise HTTPException(status_code=404, detail="Stored report file is missing")

    return FileResponse(
        report_path,
        media_type="text/csv",
        filename=report_path.name,
    )


@app.get("/api/submissions/{submission_id}")
def get_submission(submission_id: str) -> JSONResponse:
    with get_conn() as conn:
        session_row = conn.execute(
            "SELECT * FROM submission_sessions WHERE submission_id = ?",
            (submission_id,),
        ).fetchone()

        if not session_row:
            raise HTTPException(status_code=404, detail="Submission not found")

        soldier_rows = conn.execute(
            """
            SELECT name, detail_number, wbt, rir, mcs_level, mcs_shuttle
            FROM soldier_scores WHERE submission_id = ? ORDER BY rowid ASC
            """,
            (submission_id,),
        ).fetchall()

    return JSONResponse(
        {
            "submission": dict(session_row),
            "soldiers": [dict(row) for row in soldier_rows],
        }
    )
